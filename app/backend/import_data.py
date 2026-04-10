"""
Data importer — loads all CSV and Excel files into the SQLite database.
Run once (or re-run to refresh): python import_data.py
"""
import os
import sys
import csv
import json

# Allow running from any directory
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import create_app
from models.database import (
    db, RawMaterial, Product, ProductSpec, ProductRawMaterialMap,
    AlphaCode, RalPantoneShade, Stock, LabResult, ClientProductMapping
)

DATA_DIR = BASE_DIR  # CSVs are in the root workspace folder


def safe_float(val):
    try:
        if val is None or str(val).strip() == "":
            return None
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def safe_bool(val):
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() in ("true", "1", "yes")


def import_raw_materials():
    print("Importing raw_material.csv ...")
    path = os.path.join(DATA_DIR, "raw_material.csv")
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            rm = RawMaterial.query.get(row["rawmaterialid"].strip())
            if rm is None:
                rm = RawMaterial(rawmaterialid=row["rawmaterialid"].strip())
            rm.rawmaterialname = row.get("rawmaterialname", "").strip()
            rm.current_price = safe_float(row.get("current_price"))
            rm.current_price_date = row.get("current_price_date", "").strip() or None
            rm.chemical_name = row.get("chemical_name", "").strip() or None
            rm.last_price = safe_float(row.get("last_price"))
            rm.last_price_date = row.get("last_price_date", "").strip() or None
            rm.remarks = row.get("remarks", "").strip() or None
            rm.type = row.get("type", "").strip() or None
            rm.compliance = row.get("compliance", "").strip() or None
            db.session.merge(rm)
            count += 1
    db.session.commit()
    print(f"  → {count} raw materials imported.")


def import_lab_values():
    """Import LAB values from Lab_Values_Color.xlsx into RawMaterial records."""
    print("Importing Lab_Values_Color.xlsx ...")
    path = os.path.join(DATA_DIR, "Lab_Values_Color.xlsx")

    # Explicit mapping: xlsx pigment name → rawmaterialid(s) in the inventory.
    # Required because xlsx names (e.g. "RED 635") use a different convention than
    # raw_material.csv (e.g. "PR57:1-SD RD 635"), so ilike substring matching fails
    # for almost every pigment.
    PIGMENT_LAB_ID_MAP = {
        "YELLOW 3033K": ["RM992"],   # SOLVENT YELLOW 3033 K
        "YELLOW 114":   ["RM507"],   # PY62- SD YL 113/114
        "YELLOW 2909":  ["RM515"],   # PY180- SD YL 2909
        "YELLOW 2939 K": ["RM512"],  # PY139- SD YL 2939K
        "YELLOW 2925K": ["RM510"],   # PY110- SD YL 2925
        "ORANGE 2917":  ["RM705"],   # PO64-SD OR 2917
        "ORANGE 212":   ["RM702"],   # PO34-SD OR 212
        "RED 5016":     ["RM307"],   # PR53:1-SD RD 5016
        "RED 507K":     ["RM304"],   # PR53:1-SD RD 507
        "RED 570":      ["RM302"],   # PR48:3-SD RD 570
        "RED2963K":     ["RM311"],   # PR170-SD RD 2963
        "RED 2967K":    ["RM312"],   # PR170-SD RD 2967
        "RED 2985":     ["RM383"],   # RED 2985
        "RED 635":      ["RM309"],   # PR57:1-SD RD 635
        "RED 2991K":    ["RM310"],   # PR122-SD RD 2991
        "VOILET 2945K": ["RM801"],   # PV23- SD VL 2945
        "VOILET 2946K": ["RM803"],   # PV23- PD VL KBLB
        "BLUE 2749":    ["RM602"],   # PB15:1-SD BU 2749
        "BLUE 2789":    ["RM605"],   # PB15:3-SD BU 2789
        "GREEN 2730K":  ["RM401"],   # PG7-SD GR2727/2730
    }

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Lab Values"]

        # Remove stale LAB_ stub records from previous imports so phantom pigments
        # don't pollute K-M suggestions after re-import.
        stale = RawMaterial.query.filter(RawMaterial.rawmaterialid.like("LAB_%")).all()
        for s in stale:
            db.session.delete(s)
        db.session.flush()

        count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):  # skip 2 header rows
            if not row[0]:
                continue
            pigment_name = str(row[0]).strip()
            ft_L = safe_float(row[1])
            ft_a = safe_float(row[2])
            ft_b = safe_float(row[3])
            tt_L = safe_float(row[4])
            tt_a = safe_float(row[5])
            tt_b = safe_float(row[6])
            ft_hex = str(row[7]).strip() if row[7] else None
            tt_hex = str(row[8]).strip() if row[8] else None

            # Fix known data error: YELLOW 2925K tint tone a = 1293 should be 12.93
            if pigment_name == "YELLOW 2925K" and tt_a and tt_a > 100:
                tt_a = tt_a / 100.0

            def _apply_lab(rm):
                rm.full_tone_L = ft_L
                rm.full_tone_a = ft_a
                rm.full_tone_b = ft_b
                rm.tint_tone_L = tt_L
                rm.tint_tone_a = tt_a
                rm.tint_tone_b = tt_b
                rm.full_tone_hex = ft_hex
                rm.tint_tone_hex = tt_hex

            # 1. Explicit ID mapping — most reliable, checked first.
            if pigment_name in PIGMENT_LAB_ID_MAP:
                for rm_id in PIGMENT_LAB_ID_MAP[pigment_name]:
                    rm = RawMaterial.query.get(rm_id)
                    if rm:
                        _apply_lab(rm)
                        count += 1
                    else:
                        print(f"  ! Mapped ID {rm_id} not found for '{pigment_name}'")
                continue

            # 2. ilike fallback for any pigments not in the explicit map.
            rm_list = RawMaterial.query.filter(
                RawMaterial.rawmaterialname.ilike(f"%{pigment_name}%")
            ).all()
            if rm_list:
                for rm in rm_list:
                    _apply_lab(rm)
                    count += 1
                continue

            # 3. No inventory match — create a standalone LAB stub so the pigment
            #    is still available for K-M suggestions even without an inventory record.
            rm_id = f"LAB_{pigment_name.replace(' ', '_').upper()}"
            rm = RawMaterial(rawmaterialid=rm_id)
            rm.rawmaterialname = pigment_name
            rm.type = "PG"
            rm.compliance = "NON-R"
            _apply_lab(rm)
            db.session.merge(rm)
            count += 1
            print(f"  ~ No inventory match for '{pigment_name}' — created stub {rm_id}")

        db.session.commit()
        print(f"  → {count} pigment LAB entries imported.")
    except Exception as e:
        print(f"  ! Error importing Lab_Values_Color.xlsx: {e}")


def import_products():
    print("Importing products.csv ...")
    path = os.path.join(DATA_DIR, "products.csv")
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            prod = Product.query.get(str(row["id"]).strip())
            if prod is None:
                prod = Product(id=str(row["id"]).strip())
            prod.name = row.get("name", "").strip()
            prod.selling_price = safe_float(row.get("selling_price"))
            prod.remark = row.get("remark", "").strip() or None
            prod.date_updated = row.get("date_updated", "").strip() or None
            prod.alphacode = row.get("alphacode", "").strip() or None
            prod.slf_no = row.get("slf_no", "").strip() or None
            prod.ral_shade = row.get("ral_shade", "").strip() or None
            prod.pantone_shade = row.get("pantone_shade", "").strip() or None
            prod.is_final_good = safe_bool(row.get("is_final_good", "False"))
            db.session.merge(prod)
            count += 1
    db.session.commit()
    print(f"  → {count} products imported.")


def import_product_specs():
    print("Importing product_spec.csv ...")
    path = os.path.join(DATA_DIR, "product_spec.csv")
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            spec_id = safe_float(row.get("id"))
            if spec_id is None:
                continue
            spec = ProductSpec.query.get(int(spec_id))
            if spec is None:
                spec = ProductSpec(id=int(spec_id))
            spec.product_id = str(row.get("product_id", "")).strip() or None
            spec.carrier_resin = row.get("carrier_resin", "").strip() or None
            spec.material_form = row.get("material_form", "").strip() or None
            spec.appearance_colour = row.get("appearance_colour", "").strip() or None
            spec.hardness = safe_float(row.get("hardness"))
            spec.moisture_content = safe_float(row.get("moisture_content"))
            spec.antioxidant_content = safe_float(row.get("antioxidant_content"))
            spec.ash_content = safe_float(row.get("ash_content"))
            spec.specific_gravity = safe_float(row.get("specific_gravity"))
            spec.bulk_density = safe_float(row.get("bulk_density"))
            spec.caco3_content = safe_float(row.get("caco3_content"))
            spec.tio2_percentage = safe_float(row.get("tio2_percentage"))
            spec.mfr = safe_float(row.get("mfr"))
            spec.weather_fastness = safe_float(row.get("weather_fastness"))
            spec.light_fastness = safe_float(row.get("light_fastness"))
            spec.colour_migration = safe_float(row.get("colour_migration"))
            spec.dispersion = row.get("dispersion", "").strip() or None
            spec.melting_temperature = safe_float(row.get("melting_temperature"))
            spec.heat_stability = safe_float(row.get("heat_stability"))
            spec.toxicity = row.get("toxicity", "").strip() or None
            spec.spec_compliance = row.get("spec_compliance", "").strip() or None
            spec.higher_olefin_constituent = row.get("higher_olefin_constituent", "").strip() or None
            spec.cbc = row.get("cbc", "").strip() or None
            spec.carbon_type = row.get("carbon_type", "").strip() or None
            spec.volatile_matter_content = safe_float(row.get("volatile_matter_content"))
            spec.toluene_extract = safe_float(row.get("toluene_extract"))
            spec.let_down_ratio = row.get("let_down_ratio", "").strip() or None
            spec.ral_pantone_ci = row.get("ral_pantone_ci", "").strip() or None
            db.session.merge(spec)
            count += 1
    db.session.commit()
    print(f"  → {count} product specs imported.")


def import_product_raw_material_map():
    print("Importing productrawmaterialmap.csv ...")
    path = os.path.join(DATA_DIR, "productrawmaterialmap.csv")
    # Delete existing and re-import (recipes are bulk data)
    ProductRawMaterialMap.query.delete()
    db.session.commit()
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        batch = []
        for row in reader:
            productid = str(row.get("productid", "")).strip()
            rawmaterialid = str(row.get("rawmaterialid", "")).strip()
            qty = safe_float(row.get("qtyinkg"))
            if not productid or not rawmaterialid:
                continue
            batch.append(ProductRawMaterialMap(
                productid=productid,
                rawmaterialid=rawmaterialid,
                qtyinkg=qty
            ))
            if len(batch) >= 500:
                db.session.bulk_save_objects(batch)
                db.session.commit()
                batch = []
        if batch:
            db.session.bulk_save_objects(batch)
            db.session.commit()
    count = ProductRawMaterialMap.query.count()
    print(f"  → {count} recipe entries imported.")


def import_alpha_codes():
    print("Importing alphacode.csv ...")
    path = os.path.join(DATA_DIR, "alphacode.csv")
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            code = row.get("alpha_code", "").strip()
            if not code:
                continue
            ac = AlphaCode.query.get(code)
            if ac is None:
                ac = AlphaCode(alpha_code=code)
            ac.polymer = row.get("polymer", "").strip()
            ac.product_type = row.get("product_type", "").strip()
            ac.compliance = row.get("compliance", "").strip()
            ac.application = row.get("application", "").strip()
            ac.sub_application = row.get("sub_application", "").strip()
            ac.code1 = row.get("code1", "").strip()
            ac.code2 = row.get("code2", "").strip()
            ac.code3 = row.get("code3", "").strip()
            ac.code4 = row.get("code4", "").strip()
            ac.code5 = row.get("code5", "").strip()
            ac.product_code = row.get("product_code", "").strip()
            ac.gross_margin = safe_float(row.get("gross_margin"))
            db.session.merge(ac)
            count += 1
    db.session.commit()
    print(f"  → {count} alpha codes imported.")


def import_ral_pantone():
    print("Importing ral_pantone_shade.csv ...")
    path = os.path.join(DATA_DIR, "ral_pantone_shade.csv")
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            code = row.get("shade_code", "").strip()
            if not code:
                continue
            shade = RalPantoneShade.query.get(code)
            if shade is None:
                shade = RalPantoneShade(shade_code=code)
            shade.color_name = row.get("color_name", "").strip()
            shade.hex_code = row.get("hex_code", "").strip()
            db.session.merge(shade)
            count += 1
    db.session.commit()
    print(f"  → {count} RAL/Pantone shades imported.")


def import_stocks():
    print("Importing stocks.csv ...")
    path = os.path.join(DATA_DIR, "stocks.csv")
    Stock.query.delete()
    db.session.commit()
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            rm_id = row.get("rawmaterialid", "").strip() or None
            # Only link FK if the raw material actually exists
            if rm_id and not RawMaterial.query.get(rm_id):
                rm_id = None
            s = Stock(
                rawmaterialid=rm_id,
                rawmaterialname=row.get("rawmaterialname", "").strip() or None,
                available_stocks=safe_float(row.get("available_stocks")),
                particulars_name=row.get("particulars_name", "").strip() or None,
                last_updated=row.get("last_updated", "").strip() or None,
            )
            db.session.add(s)
            count += 1
    db.session.commit()
    print(f"  → {count} stock entries imported.")


def import_client_product_mapping():
    print("Importing clientproductmapping.csv ...")
    path = os.path.join(DATA_DIR, "clientproductmapping.csv")
    ClientProductMapping.query.delete()
    db.session.commit()
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            c = ClientProductMapping(
                client_id=row.get("client_id", "").strip(),
                product_id=row.get("product_id", "").strip() or None,
                premium_disc=safe_float(row.get("premium_disc")),
                alphacode=row.get("alphacode", "").strip() or None,
            )
            db.session.add(c)
            count += 1
    db.session.commit()
    print(f"  → {count} client-product mappings imported.")


def import_achromatic_lab():
    """
    Inject well-known LAB values for carbon blacks and TiO2 whites.
    These pigments are not in Lab_Values_Color.xlsx but are essential for
    grey/black/white K-M suggestions.

    Full-tone: the pigment at 100% loading.
    Tint-tone: pigment:TiO2 = 1:10 (same convention as the xlsx).
    Values are typical industry references for these pigment grades.
    """
    ACHROMATIC_DATA = [
        # (rawmaterialid, full_L, full_a, full_b, tint_L, tint_a, tint_b)
        ("RM181", 4.5, 0.5, 0.5, 44.0, 0.3, 0.5),   # CARBON 660 (standard channel black)
        ("RM183", 4.0, 0.5, 0.8, 40.0, 0.4, 0.6),   # CARBON 330 (furnace black)
        ("RM187", 3.5, 0.4, 0.4, 37.0, 0.3, 0.4),   # CARBON HG (high jetness)
        ("RM202", 99.0, 0.0, 1.5, 99.0, 0.0, 1.5),  # TT 100 (rutile TiO2)
        ("RM204", 99.0, -0.1, 1.8, 99.0, -0.1, 1.8), # TT 666 (rutile TiO2)
    ]
    count = 0
    for rm_id, ft_L, ft_a, ft_b, tt_L, tt_a, tt_b in ACHROMATIC_DATA:
        rm = RawMaterial.query.get(rm_id)
        if rm is None:
            print(f"  ! Achromatic RM {rm_id} not found — skipped.")
            continue
        rm.full_tone_L = ft_L
        rm.full_tone_a = ft_a
        rm.full_tone_b = ft_b
        rm.tint_tone_L = tt_L
        rm.tint_tone_a = tt_a
        rm.tint_tone_b = tt_b
        count += 1
    db.session.commit()
    print(f"  → {count} achromatic (carbon/TiO2) LAB values set.")


def import_pigment_properties():
    """Import pigment technical properties from pigment_properties.csv.
    Data sourced from Sudarshan shade card PDFs (1.Plastic Shadecard.pdf,
    2.Secondery Plastic.pdf): CI name, chemistry, light/weather fastness,
    heat resistance, bleed/warp behaviour.
    Must run after import_lab_values() so LAB_ stubs are available.
    """
    print("Importing pigment_properties.csv ...")
    path = os.path.join(DATA_DIR, "pigment_properties.csv")
    if not os.path.exists(path):
        print("  ! pigment_properties.csv not found — skipped.")
        return

    # Explicit pigment name → rawmaterialid mapping
    PROPS_ID_MAP = {
        "YELLOW 114":    "RM507",   "YELLOW 132K":   "RM517",
        "YELLOW 2909":   "RM515",   "YELLOW 2939 K": "RM512",
        "YELLOW 2925K":  "RM510",   "ORANGE 2917":   "RM705",
        "ORANGE 212":    "RM702",   "RED 5016":      "RM307",
        "RED 570":       "RM302",   "RED 2967K":     "RM312",
        "RED 2985":      "RM383",   "RED 635":       "RM309",
        "RED 2991K":     "RM310",   "BLUE 2789":     "RM605",
        "GREEN 2730K":   "RM401",   "YELLOW 162K":   "RM503",
        "YELLOW 137K":   "RM508",   "ORANGE 203K":   "RM701",
        "RED 507K":      "RM304",   "RED 2957K":     "RM313",
        "RED 2963K":     "RM311",   "RED 565":       "RM300",
        "RED 587":       "RM308",   "BLUE 2633K":    "RM600",
    }

    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        skipped = []
        for row in reader:
            name = row["pigment_name"].strip()
            rm = None
            if name in PROPS_ID_MAP:
                rm = RawMaterial.query.get(PROPS_ID_MAP[name])
            if rm is None:
                # Fallback: LAB_ stub created by import_lab_values for unmatched pigments
                stub_id = f"LAB_{name.replace(' ', '_').upper()}"
                rm = RawMaterial.query.get(stub_id)
            if rm is None:
                skipped.append(name)
                continue
            rm.ci_name = row["ci_name"].strip() or None
            rm.chemistry = row["chemistry"].strip() or None
            rm.heat_resistance = safe_float(row.get("heat_resistance"))
            rm.light_fastness_tone = safe_float(row.get("light_fastness_tone"))
            rm.light_fastness_tint = safe_float(row.get("light_fastness_tint"))
            rm.weather_fastness_tone = safe_float(row.get("weather_fastness_tone"))
            rm.weather_fastness_tint = safe_float(row.get("weather_fastness_tint"))
            rm.bleed_pvc = row.get("bleed_pvc", "").strip() or None
            rm.warp_hdpe = row.get("warp_hdpe", "").strip() or None
            count += 1
    db.session.commit()
    if skipped:
        print(f"  ~ Not matched (no RM or stub): {skipped}")
    print(f"  → {count} pigment records updated with technical properties.")


def import_fgdata_lab():
    """Import spectrophotometer L/a/b readings from lab_results.csv.
    Extracted from FG MasterData FGData sheet (235 products measured by spectrophotometer).
    Must run after import_products() so FK constraints are satisfied.
    """
    print("Importing lab_results.csv (FGData spectro readings) ...")
    path = os.path.join(DATA_DIR, "lab_results.csv")
    if not os.path.exists(path):
        print("  ! lab_results.csv not found — skipped.")
        return

    # Clear previous FGData-sourced results to allow clean re-import
    deleted = LabResult.query.filter(LabResult.notes == "FGData").delete()
    db.session.commit()
    if deleted:
        print(f"  ~ Cleared {deleted} previous FGData lab results.")

    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        skipped = 0
        for row in reader:
            pid = row["product_id"].strip()
            L = safe_float(row.get("L"))
            a = safe_float(row.get("a"))
            b = safe_float(row.get("b"))
            if L is None or a is None or b is None:
                skipped += 1
                continue
            polymer = row.get("polymer", "PE").strip() or "PE"
            lr = LabResult(
                product_id=pid,
                polymer=polymer,
                L=L,
                a=a,
                b=b,
                notes="FGData",
            )
            db.session.add(lr)
            count += 1
            if count % 100 == 0:
                db.session.flush()

    db.session.commit()
    if skipped:
        print(f"  ~ {skipped} rows skipped (null values or product not in DB).")
    print(f"  → {count} FGData lab results imported.")


def import_lab_values_assumed():
    """Import assumed / average LAB reference values for carrier resins, additives,
    and fillers from LAB_values_assumed.xlsx.

    Values are indicative (not instrument-tested) and are stored in the
    full_tone_L / full_tone_a / full_tone_b fields.  Only empty fields are
    written so that actual spectrophotometer readings are never overwritten.
    """
    print("Importing LAB_values_assumed.xlsx (assumed raw-material colours) ...")
    path = os.path.join(DATA_DIR, "LAB_values_assumed.xlsx")
    if not os.path.exists(path):
        print("  ! LAB_values_assumed.xlsx not found — skipped.")
        return

    # Explicit overrides for names that differ from the inventory.
    EXPLICIT_MAP = {
        "RESIN":           "RM040",   # PVC K67  (generic PVC resin)
        "RESIN K70":       "RM041",   # PVC K70
        "PE Wax - CL-60":  "RM113",  # PE WAX CL-60
    }

    # Names confirmed absent from inventory — skip silently.
    SKIP_NAMES = {"LASER 8825", "VCI RESICOR-002", "ZINC STEARATE"}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["LAB Values"]

        # Build a quick lookup: lowercase name → RawMaterial (RM-type only)
        rm_by_name = {}
        for rm in RawMaterial.query.filter(RawMaterial.type == "RM").all():
            if rm.rawmaterialname:
                rm_by_name[rm.rawmaterialname.strip().lower()] = rm

        count = 0
        skipped = []

        # Data starts at row 5 (rows 1-4 are title / description / blank / header)
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row[0]:
                continue
            mat_name = str(row[0]).strip()
            L = safe_float(row[1])
            a = safe_float(row[2])
            b = safe_float(row[3])
            if L is None or a is None or b is None:
                continue
            if mat_name in SKIP_NAMES:
                continue

            rm = None
            if mat_name in EXPLICIT_MAP:
                rm = RawMaterial.query.get(EXPLICIT_MAP[mat_name])
            else:
                # Exact case-insensitive match first
                rm = rm_by_name.get(mat_name.lower())

            if rm is None:
                # Substring fallback: xlsx name contained in or containing db name
                matches = [v for k, v in rm_by_name.items()
                           if mat_name.lower() in k or k in mat_name.lower()]
                if len(matches) == 1:
                    rm = matches[0]

            if rm is None:
                skipped.append(mat_name)
                continue

            # Only set if not already populated (preserve actual readings)
            if rm.full_tone_L is None:
                rm.full_tone_L = L
                rm.full_tone_a = a
                rm.full_tone_b = b
                count += 1

        db.session.commit()
        if skipped:
            print(f"  ~ No inventory match for: {skipped}")
        print(f"  → {count} raw-material assumed LAB values set.")

    except Exception as e:
        print(f"  ! Error importing LAB_values_assumed.xlsx: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    app = create_app()
    with app.app_context():
        db.create_all()
        print("Database tables created.\n")
        import_raw_materials()
        import_lab_values()
        import_achromatic_lab()
        import_pigment_properties()
        import_lab_values_assumed()
        import_products()
        import_product_specs()
        import_product_raw_material_map()
        import_alpha_codes()
        import_ral_pantone()
        import_stocks()
        import_client_product_mapping()
        import_fgdata_lab()
        print("\nAll data imported successfully.")
